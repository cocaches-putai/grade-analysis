# src/exporter.py
from typing import Dict, List, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.stats import get_subject_cols, sort_grades


# ── 基礎工具 ──────────────────────────────────────────────────────
def export_to_excel(sheets: Dict[str, pd.DataFrame], output_path: str) -> None:
    """將多個 DataFrame 匯出為多頁 Excel 檔案（通用版）。"""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)


# ── 班別排序 ──────────────────────────────────────────────────────
_CLASS_SUFFIX_ORDER = ["甲", "乙", "丙", "丁", "戊", "己", "庚", "辛", "壬", "癸"]

def _sort_classes(classes) -> list:
    def key(c):
        for i, ch in enumerate(_CLASS_SUFFIX_ORDER):
            if c.endswith(ch):
                return i
        return 99
    return sorted(classes, key=key)


# ── 類組判斷 ──────────────────────────────────────────────────────
def get_class_track(class_name: str) -> str:
    """
    高一 → 普通科
    高二/高三 甲乙 → 社會組
    高二/高三 己庚 → 自然組
    """
    if "高一" in class_name:
        return "普通科"
    if any(g in class_name for g in ("高二", "高三")):
        if class_name.endswith(("甲", "乙")):
            return "社會組"
        if class_name.endswith(("己", "庚")):
            return "自然組"
    return "其他"


# ── 樣式 ──────────────────────────────────────────────────────────
_HEADER_FILL  = PatternFill("solid", fgColor="D9E1F2")
_SUBJ_FILL    = PatternFill("solid", fgColor="BDD7EE")
_AVG_FILL     = PatternFill("solid", fgColor="E2EFDA")
_TEACHER_FILL = PatternFill("solid", fgColor="FFF2CC")
_TITLE_FILL   = PatternFill("solid", fgColor="F2F2F2")
_THIN         = Side(style="thin")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

def _cell(ws, row, col, value=None, bold=False, fill=None,
          align="center", num_fmt=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = _BORDER
    if bold:
        c.font = Font(bold=True)
    if fill:
        c.fill = fill
    if num_fmt:
        c.number_format = num_fmt
    return c

def _title_row(ws, row, text, total_cols):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=12)
    c.fill = _TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    if total_cols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=total_cols)


# ── 年級各科分頁 ──────────────────────────────────────────────────
_BANDS = [
    ("90-100", 90, 101), ("80-89", 80, 90), ("70-79", 70, 80),
    ("60-69", 60, 70), ("50-59", 50, 60), ("40-49", 40, 50),
    ("30-39", 30, 40), ("20-29", 20, 30), ("10-19", 10, 20), ("0-9", 0, 10),
]

def _write_grade_sheet(wb, grade: str, grade_df: pd.DataFrame,
                       subjects: List[str], teacher_map: dict,
                       school_name: str, exam_name: str):
    ws = wb.create_sheet(title=f"{grade}各科")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 10

    # 標題列
    ws.row_dimensions[1].height = 30
    _title_row(ws, 1, f"{school_name}　{grade}　{exam_name}　各科分組/班級成績分析", 8)
    current_row = 2

    for subj in subjects:
        valid_classes = _sort_classes([
            cls for cls in grade_df["班級"].unique()
            if len(grade_df[grade_df["班級"] == cls][subj].dropna()) > 0
        ])
        if not valid_classes:
            continue

        nc = len(valid_classes)

        # 科目標題（跨欄）
        _cell(ws, current_row, 2, "科目", bold=True, fill=_SUBJ_FILL)
        _cell(ws, current_row, 3, subj, bold=True, fill=_SUBJ_FILL)
        if nc > 1:
            ws.merge_cells(start_row=current_row, start_column=3,
                           end_row=current_row, end_column=2 + nc)
        current_row += 1

        # 班級
        _cell(ws, current_row, 2, "班級", bold=True, fill=_HEADER_FILL)
        for i, cls in enumerate(valid_classes):
            ws.column_dimensions[get_column_letter(3 + i)].width = 8
            _cell(ws, current_row, 3 + i, cls, bold=True, fill=_HEADER_FILL)
        current_row += 1

        # 到考人數
        _cell(ws, current_row, 2, "到考人數", fill=_HEADER_FILL)
        for i, cls in enumerate(valid_classes):
            n = len(grade_df[grade_df["班級"] == cls][subj].dropna())
            _cell(ws, current_row, 3 + i, n)
        current_row += 1

        # 不及格人數
        _cell(ws, current_row, 2, "不及格人數", fill=_HEADER_FILL)
        for i, cls in enumerate(valid_classes):
            subset = grade_df[grade_df["班級"] == cls][subj].dropna()
            _cell(ws, current_row, 3 + i, int((subset < 60).sum()))
        current_row += 1

        # 分數區間
        for label, lo, hi in _BANDS:
            _cell(ws, current_row, 2, label)
            for i, cls in enumerate(valid_classes):
                subset = grade_df[grade_df["班級"] == cls][subj].dropna()
                _cell(ws, current_row, 3 + i,
                      int(((subset >= lo) & (subset < hi)).sum()))
            current_row += 1

        # 平均
        _cell(ws, current_row, 2, "平均", bold=True, fill=_AVG_FILL)
        for i, cls in enumerate(valid_classes):
            subset = grade_df[grade_df["班級"] == cls][subj].dropna()
            val = round(float(subset.mean()), 2) if len(subset) > 0 else None
            _cell(ws, current_row, 3 + i, val, fill=_AVG_FILL, num_fmt="0.00")
        current_row += 1

        # 任課老師
        _cell(ws, current_row, 2, "任課老師", fill=_TEACHER_FILL)
        for i, cls in enumerate(valid_classes):
            teacher = teacher_map.get(subj, {}).get(cls, "")
            _cell(ws, current_row, 3 + i, teacher or None, fill=_TEACHER_FILL)
        current_row += 1

        current_row += 3  # 科目間空行


# ── 各科平均總表（排除英數）────────────────────────────────────────
_ENG_MATH_SUBJECTS = {"英語文", "數學"}

def _write_summary_sheet(wb, df: pd.DataFrame, subjects: List[str],
                         teacher_map: dict, exam_name: str,
                         all_classes: List[str], school_name: str):
    # 排除英語文/數學（含所有變體），排除選修課
    _EXCLUDE_KW = ("英語", "英文", "數學", "選修", "進階", "閱讀", "作文",
                   "英聽", "空間資訊", "現代社會", "思考：", "探究與實作")
    non_eng_math = [
        s for s in subjects
        if not any(kw in s for kw in _EXCLUDE_KW)
        and df[s].notna().any()
    ]
    if not non_eng_math:
        return

    ws = wb.create_sheet(title="各科平均總表")
    total_cols = 2 + len(non_eng_math)

    # 標題
    ws.row_dimensions[1].height = 24
    _title_row(ws, 1, f"{school_name}　{exam_name}　各科班級平均成績", total_cols)

    # 欄頭
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    _cell(ws, 2, 1, "班級", bold=True, fill=_HEADER_FILL)
    _cell(ws, 2, 2, "科目", bold=True, fill=_HEADER_FILL)
    for i, subj in enumerate(non_eng_math):
        ws.column_dimensions[get_column_letter(3 + i)].width = 10
        _cell(ws, 2, 3 + i, subj, bold=True, fill=_HEADER_FILL)

    row = 3
    for cls in all_classes:
        # 任課老師列
        _cell(ws, row, 1, cls, bold=True, fill=_TEACHER_FILL)
        _cell(ws, row, 2, "任課老師", fill=_TEACHER_FILL)
        for i, subj in enumerate(non_eng_math):
            teacher = teacher_map.get(subj, {}).get(cls, "")
            _cell(ws, row, 3 + i, teacher or None, fill=_TEACHER_FILL)
        row += 1

        # 成績列
        _cell(ws, row, 1, "", fill=_AVG_FILL)
        _cell(ws, row, 2, exam_name, fill=_AVG_FILL)
        for i, subj in enumerate(non_eng_math):
            subset = df[df["班級"] == cls][subj].dropna()
            val = round(float(subset.mean()), 2) if len(subset) > 0 else None
            _cell(ws, row, 3 + i, val, fill=_AVG_FILL, num_fmt="0.00")
        row += 1


# ── 英數平均總表 ──────────────────────────────────────────────────
_ENG_MATH_KW = ("英語", "英文", "數學", "英聽")

def _write_eng_math_sheet(wb, df: pd.DataFrame,
                          subjects: List[str], teacher_map: dict,
                          exam_name: str, all_classes: List[str],
                          school_name: str, subject_groups: Optional[dict] = None):
    """
    subject_groups 格式（來自 grouping_parser.parse_grouping_excel）：
    {"數學": {"高一": [{"label": ..., "teacher": ..., "students": [...]}, ...], ...}, ...}
    有分組資料則依分組顯示；否則 fallback 為班別。
    """
    # 只保留主要英文/數學（排除純選修）
    _ELECTIVE_KW = ("選修", "進階", "閱讀", "作文", "空間")
    eng_math = [
        s for s in subjects
        if any(kw in s for kw in _ENG_MATH_KW)
        and not any(kw in s for kw in _ELECTIVE_KW)
        and df[s].notna().any()
    ]
    if not eng_math:
        return

    ws = wb.create_sheet(title="英數平均總表")
    total_cols = 2 + len(eng_math)

    ws.row_dimensions[1].height = 24
    _title_row(ws, 1, f"{school_name}　{exam_name}　英數分組平均成績", total_cols)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    _cell(ws, 2, 1, "分組/班級", bold=True, fill=_HEADER_FILL)
    _cell(ws, 2, 2, "科目", bold=True, fill=_HEADER_FILL)
    for i, subj in enumerate(eng_math):
        ws.column_dimensions[get_column_letter(3 + i)].width = 10
        _cell(ws, 2, 3 + i, subj, bold=True, fill=_HEADER_FILL)

    row = 3
    grades = sort_grades(df["年級"].unique().tolist())

    for grade in grades:
        grade_classes = _sort_classes(
            df[df["年級"] == grade]["班級"].unique().tolist()
        )

        # 決定使用分組模式還是班別 fallback
        # 分組模式：該年段有任一科目有分組資料
        use_groups = False
        grade_groups_by_subj: dict = {}
        if subject_groups:
            for subj in eng_math:
                if subj in subject_groups and grade in subject_groups[subj]:
                    grade_groups_by_subj[subj] = subject_groups[subj][grade]
                    use_groups = True

        if use_groups:
            # 以「有分組的科目」的組數為基準（取最多組的那個科目）
            ref_subj = max(grade_groups_by_subj, key=lambda s: len(grade_groups_by_subj[s]))
            ref_groups = grade_groups_by_subj[ref_subj]

            for grp in ref_groups:
                student_set = set(grp["students"])
                label = grp["label"]

                # 老師：優先用分組資料，不足的科目用 teacher_map
                _cell(ws, row, 1, label, bold=True, fill=_TEACHER_FILL)
                _cell(ws, row, 2, "任課老師", fill=_TEACHER_FILL)
                for i, subj in enumerate(eng_math):
                    # 從分組資料找這個 label 的老師
                    teacher = ""
                    if subj in grade_groups_by_subj:
                        matched = next(
                            (g for g in grade_groups_by_subj[subj]
                             if g["label"] == label or g["code"] == grp["code"]),
                            None
                        )
                        if matched:
                            teacher = matched["teacher"]
                    _cell(ws, row, 3 + i, teacher or None, fill=_TEACHER_FILL)
                row += 1

                # 成績：從 scores_df 找到分組學生的成績
                _cell(ws, row, 1, "", fill=_AVG_FILL)
                _cell(ws, row, 2, exam_name, fill=_AVG_FILL)
                for i, subj in enumerate(eng_math):
                    # 如有此科的分組，用分組學生清單；否則用整個年段
                    if subj in grade_groups_by_subj:
                        s_set = set(next(
                            (g["students"] for g in grade_groups_by_subj[subj]
                             if g["label"] == label or g["code"] == grp["code"]),
                            grp["students"]
                        ))
                    else:
                        s_set = student_set
                    subset = df[df["姓名"].isin(s_set)][subj].dropna()
                    val = round(float(subset.mean()), 2) if len(subset) > 0 else None
                    _cell(ws, row, 3 + i, val, fill=_AVG_FILL, num_fmt="0.00")
                row += 1

        else:
            # fallback：依班別顯示
            for cls in grade_classes:
                has_data = any(
                    len(df[df["班級"] == cls][s].dropna()) > 0 for s in eng_math
                )
                if not has_data:
                    continue

                _cell(ws, row, 1, cls, bold=True, fill=_TEACHER_FILL)
                _cell(ws, row, 2, "任課老師", fill=_TEACHER_FILL)
                for i, subj in enumerate(eng_math):
                    teacher = teacher_map.get(subj, {}).get(cls, "")
                    _cell(ws, row, 3 + i, teacher or None, fill=_TEACHER_FILL)
                row += 1

                _cell(ws, row, 1, "", fill=_AVG_FILL)
                _cell(ws, row, 2, exam_name, fill=_AVG_FILL)
                for i, subj in enumerate(eng_math):
                    subset = df[df["班級"] == cls][subj].dropna()
                    val = round(float(subset.mean()), 2) if len(subset) > 0 else None
                    _cell(ws, row, 3 + i, val, fill=_AVG_FILL, num_fmt="0.00")
                row += 1


# ── 前三名 ────────────────────────────────────────────────────────
def _write_top3_sheet(wb, df: pd.DataFrame, subjects: List[str],
                      all_classes: List[str], school_name: str, exam_name: str):
    ws = wb.create_sheet(title="各班前三名")

    headers = ["類組", "班級", "姓名", "班名次", "類組名次", "年級名次", "平均分"]
    col_widths = [8, 8, 8, 7, 7, 7, 8]
    for i, (h, w) in enumerate(zip(headers, col_widths)):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
        _cell(ws, 2, i + 1, h, bold=True, fill=_HEADER_FILL)

    ws.row_dimensions[1].height = 24
    _title_row(ws, 1, f"{school_name}　{exam_name}　各班前三名", len(headers))

    # 計算每位學生的平均分
    working = df.copy()
    avail_subjs = [s for s in subjects if working[s].notna().any()]
    working["_avg"] = working[avail_subjs].mean(axis=1)

    # 班名次
    for cls in all_classes:
        mask = working["班級"] == cls
        working.loc[mask, "_班名次"] = (
            working.loc[mask, "_avg"].rank(ascending=False, method="min")
        )

    # 類組名次：同一年段同一類組內排名
    working["_類組"] = working["班級"].apply(get_class_track)
    working["_年段"] = working["班級"].apply(
        lambda c: next((g for g in ("高一", "高二", "高三", "國一", "國二", "國三")
                        if c.startswith(g)), "")
    )
    for (yr, track), grp in working.groupby(["_年段", "_類組"]):
        working.loc[grp.index, "_類組名次"] = (
            grp["_avg"].rank(ascending=False, method="min")
        )

    # 年級名次：整個年段排名
    for yr, grp in working.groupby("_年段"):
        working.loc[grp.index, "_年級名次"] = (
            grp["_avg"].rank(ascending=False, method="min")
        )

    row = 3
    for cls in all_classes:
        class_df = working[working["班級"] == cls].copy()
        if class_df.empty:
            continue
        top3 = class_df.nsmallest(3, "_班名次")

        for _, r in top3.iterrows():
            _cell(ws, row, 1, get_class_track(cls))
            _cell(ws, row, 2, cls)
            _cell(ws, row, 3, r["姓名"])
            _cell(ws, row, 4, int(r["_班名次"]))
            _cell(ws, row, 5, int(r["_類組名次"]))
            _cell(ws, row, 6, int(r["_年級名次"]))
            _cell(ws, row, 7, round(float(r["_avg"]), 2), num_fmt="0.00")
            row += 1

        # 班與班間空一行
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col)
        row += 1


# ── 各類組排行榜 ──────────────────────────────────────────────────
def _write_ranking_sheet(wb, df: pd.DataFrame, subjects: List[str],
                         grades: List[str], school_name: str, exam_name: str):
    ws = wb.create_sheet(title="各類組排行榜")

    ws.row_dimensions[1].height = 24
    _title_row(ws, 1, f"{school_name}　{exam_name}　各類組排行榜", 8)

    avail_subjs = [s for s in subjects if df[s].notna().any()]
    df = df.copy()
    df["_avg"] = df[avail_subjs].mean(axis=1)
    df["_類組"] = df["班級"].apply(get_class_track)
    df["_年段"] = df["班級"].apply(
        lambda c: next((g for g in ("高一", "高二", "高三", "國一", "國二", "國三")
                        if c.startswith(g)), "")
    )

    current_col = 1
    for grade in grades:
        grade_df = df[df["_年段"] == grade]
        tracks = sorted(grade_df["_類組"].unique())

        for track in tracks:
            track_df = (
                grade_df[grade_df["_類組"] == track]
                .copy()
                .sort_values("_avg", ascending=False)
                .reset_index(drop=True)
            )
            track_df["_名次"] = track_df["_avg"].rank(
                ascending=False, method="min").astype(int)

            # 欄標題
            ws.column_dimensions[get_column_letter(current_col)].width = 6
            ws.column_dimensions[get_column_letter(current_col + 1)].width = 8
            ws.column_dimensions[get_column_letter(current_col + 2)].width = 7
            _cell(ws, 2, current_col,     f"{grade}　{track}", bold=True, fill=_SUBJ_FILL)
            ws.merge_cells(start_row=2, start_column=current_col,
                           end_row=2, end_column=current_col + 2)
            _cell(ws, 3, current_col,     "班級", bold=True, fill=_HEADER_FILL)
            _cell(ws, 3, current_col + 1, "姓名", bold=True, fill=_HEADER_FILL)
            _cell(ws, 3, current_col + 2, "名次", bold=True, fill=_HEADER_FILL)

            for i, (_, r) in enumerate(track_df.iterrows()):
                row = 4 + i
                _cell(ws, row, current_col,     r["班級"])
                _cell(ws, row, current_col + 1, r["姓名"])
                _cell(ws, row, current_col + 2, int(r["_名次"]))

            current_col += 4  # 類組間空一欄


# ── 主進入點 ──────────────────────────────────────────────────────
def export_analysis_excel(
    df: pd.DataFrame,
    teacher_map: dict,
    exam_name: str,
    output_path: str,
    school_name: str = "普台高級中學",
    subject_groups: Optional[dict] = None,
) -> None:
    """
    產生完整成績分析 Excel，格式對應學校現有手動分析檔。

    分頁：
    - 高一/二/三各科（各年級各科分組/班級成績分析）
    - 各科平均總表（排除英數）
    - 英數平均總表
    - 各班前三名（含類組名次、年級名次）
    - 各類組排行榜
    """
    subjects = get_subject_cols(df)
    grades = sort_grades(df["年級"].unique().tolist())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 各年級各科
    for grade in grades:
        grade_df = df[df["年級"] == grade]
        _write_grade_sheet(wb, grade, grade_df, subjects, teacher_map,
                           school_name, exam_name)

    # 所有班級排序
    all_classes: List[str] = []
    for grade in grades:
        all_classes.extend(_sort_classes(
            df[df["年級"] == grade]["班級"].unique().tolist()
        ))

    _write_summary_sheet(wb, df, subjects, teacher_map, exam_name,
                         all_classes, school_name)
    _write_eng_math_sheet(wb, df, subjects, teacher_map, exam_name,
                          all_classes, school_name, subject_groups=subject_groups)
    _write_top3_sheet(wb, df, subjects, all_classes, school_name, exam_name)
    _write_ranking_sheet(wb, df, subjects, grades, school_name, exam_name)

    wb.save(output_path)
